import os
import sys
import subprocess

# Ensure openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[INFO] openpyxl가 설치되어 있지 않습니다. 자동 설치를 진행합니다...")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        print("[INFO] openpyxl 설치가 완료되었습니다.")
    except Exception as e:
        print(f"[ERROR] openpyxl 설치에 실패했습니다. 수동으로 'pip install openpyxl'을 실행해 주세요. 에러: {e}")
        sys.exit(1)

def create_excel():
    # 12개 파이썬 소스 분석 데이터 정의
    data = [
        {
            "no": 1,
            "program": "p190872_local_chk_v04_gm.py",
            "params": "<스키마.검색기준테이블>\n[--db]\n[--conf <경로>]",
            "param_desc": "- <스키마.검색기준테이블>: DB 내 검색 대상 기준 테이블 (필수)\n- --db: 분석 결과 DB 테이블 적재 활성화 여부 (선택)\n- --conf: DB 접속 정보가 담긴 설정 파일 경로 (선택)",
            "opt_desc": "- --db: 지정 시 cols 파싱 결과, query_text, 매칭 결과를 DB 테이블에 자동 적재\n- --conf: 기본 mysql.conf 파일의 경로를 재지정할 때 사용",
            "features": "DB 내 검색기준테이블을 조회하여 지정된 소스 파일(source_file)의 쿼리를 추출(주석 제거 포함)하고, cols 항목의 칼럼들이 소스 내에 매칭(대소문자 무시)되는지 분석하여 3개의 결과 파일(cols 파싱, query_text, 매칭결과)을 생성하고 DB에 적재함."
        },
        {
            "no": 2,
            "program": "p190872_local_chk_v05_gm.py",
            "params": "<스키마.검색기준테이블>\n[--db]\n[--conf <경로>]",
            "param_desc": "(v04와 동일)\n- <스키마.검색기준테이블>: DB 검색 기준 테이블 (필수)\n- --db: DB 적재 활성화 여부 (선택)\n- --conf: mysql.conf 경로 (선택)",
            "opt_desc": "(v04와 동일)\n- --db: DB 적재 활성화\n- --conf: 설정 파일 경로 재정의",
            "features": "v04_gm의 확장 버전으로, 칼럼 매칭 시 오탐을 줄이기 위해 기존의 단순 포함(in 연산) 방식을 정규식 단어 완전일치(\\b단어\\b) 조건으로 수정하였으며, 결과 테이블명에 프로그램명과 구분자(_v05)가 포함되도록 명명 규칙을 보완함."
        },
        {
            "no": 3,
            "program": "p190872_local_chk_v06_gm.py",
            "params": "<검색기준테이블>\n<검색디렉토리>\n[--mid <MID목록>]\n[--db]\n[--conf <경로>]\n[--where <old/new>]\n[--chk <default/encdec_no>]",
            "param_desc": "- <검색기준테이블>: DB 검색 기준 테이블 (필수)\n- <검색디렉토리>: 소스 파일 탐색 로컬 디렉토리 (필수)\n- --mid: 검색할 하위 MID 디렉토리 목록 (쉼표 구분)\n- --where: 검색기준테이블 조회 조건 필터\n- --chk: 암호화/복호화 포함/제외 필터",
            "opt_desc": "- --mid: 미지정 시 전체 MID를 분석\n- --where: 'old' 또는 'new' 중 선택하여 기준 테이블 조회 필터링\n- --chk: 'default'(암복호화 포함) 또는 'encdec_no'(암복호화 제외) 중 분석 조건 선택",
            "features": "지정한 로컬 디렉토리 하위의 소스 파일들을 분석하여 칼럼 사용 여부를 상세 매칭함. SQL 주석 제거, Pure Column 제외 필터(Omit/Include 룰)를 적용하여 가공이나 함수가 적용된 식만 추출하며, MID별 개별 탐색, 화면 출력용 print.txt 및 제외 행 로그용 exclude.txt 파일을 생성함."
        },
        {
            "no": 4,
            "program": "p190872_local_chk_v08_gm.py",
            "params": "<검색기준테이블>\n<검색디렉토리>\n<검색결과테이블명>\n[--mid <MID목록>]\n[--db]\n[--conf <경로>]\n[--where <old/new>]\n[--chk <default/encdec_no/all>]",
            "param_desc": "- <검색결과테이블명>: 분석 결과를 적재할 DB 테이블명 (필수)\n- --chk: 암복호화 필터 조건 ('all' 조건 추가)\n(나머지는 v06과 동일)",
            "opt_desc": "- --chk: 'all' 지정 시 default와 encdec_no의 분석 결과를 각각 분리 생성하여 적재\n- <검색결과테이블명>: 대상 결과 테이블명을 파라미터로 직접 지정",
            "features": "v07_gm(주석 원본 유지, exclude 테이블 적재 기능 등)을 계승하고 암호화 검증 기능이 추가된 최종 버전. default 분리 CSV 파일 내 동일 라인에 [컬럼명], [암복호화 함수], [tobe_enc_key 변환코드(e1~e4)]의 동시 존재 여부를 검사하여 OK/NOT OK 판정 및 검증 결과 CSV를 생성함."
        },
        {
            "no": 5,
            "program": "sql_est_010_load.py",
            "params": "<소스분석할 디렉토리경로>\n[--conf <경로>]",
            "param_desc": "- <소스분석할 디렉토리경로>: 소스 파일(.sh, .hql, .sql 등)이 위치한 로컬 디렉토리 (필수)\n- --conf: mysql.conf 설정 파일 경로 (선택)",
            "opt_desc": "- --conf: DB 접속용 설정 파일 경로 지정 (미지정 시 실행 디렉토리에서 자동 탐색)",
            "features": "지정한 소스 디렉토리 내 파일들로부터 SQL 쿼리문들을 추출하여 out/ 폴더에 .dat 파일로 생성한 후, 이를 읽어 MySQL DB 테이블에 즉시 적재(Load) 처리함."
        },
        {
            "no": 6,
            "program": "sql_est_010_load_line.py",
            "params": "<소스분석할 디렉토리경로>\n[--conf <경로>]",
            "param_desc": "(sql_est_010_load와 동일)\n- <소스분석할 디렉토리경로>: 소스 파일 디렉토리\n- --conf: mysql.conf 경로",
            "opt_desc": "(sql_est_010_load와 동일)",
            "features": "sql_est_010_load.py와 기능이 동일하나, 분석 결과의 정밀도를 위해 소스 파일 내 각 쿼리가 시작되는 시작 라인 넘버(Line Number)를 함께 추출하여 결과 파일(.dat) 및 DB 적재 테이블에 기록함."
        },
        {
            "no": 7,
            "program": "sql_lng_015_emrput.py",
            "params": "<소스디렉토리>\n[--mode <SIMPLE/DETAIL>]\n[--db]\n[--conf <경로>]",
            "param_desc": "- <소스디렉토리>: SQL 리니지 분석 대상 디렉토리 (필수)\n- --mode: 리니지 분석 모드 (SIMPLE 또는 DETAIL)\n- --db: DB 적재 활성화 여부\n- --conf: mysql.conf 경로",
            "opt_desc": "- --mode: SIMPLE(기본, CTE 투명 처리 및 물리소스->타겟 추출) 또는 DETAIL(WITH절 CTE 임시테이블 흐름 포함) 중 선택\n- --db: 활성화 시 리니지 분석 결과를 DB에 자동 적재",
            "features": "소스 코드로부터 테이블 간 데이터 흐름(Lineage)을 분석하여 CSV 및 DB에 저장하며, 추가적으로 파일 내에 'EMRPUT' 문자열이 포함된 라인을 추출하여 별도의 CSV 파일로 생성하는 특화 기능이 탑재됨."
        },
        {
            "no": 8,
            "program": "sql_lng_015_with.py",
            "params": "<소스디렉토리>\n[--mode <SIMPLE/DETAIL>]\n[--db]\n[--conf <경로>]",
            "param_desc": "(sql_lng_015_emrput과 동일)\n- <소스디렉토리>: 분석 대상 디렉토리\n- --mode: 분석 모드 (SIMPLE 또는 DETAIL)",
            "opt_desc": "(sql_lng_015_emrput과 동일)",
            "features": "SQL 리니지 분석 및 DB/CSV 저장 도구로, 특히 WITH 절로 시작하는 CTE(Common Table Expression) 쿼리의 흐름을 파싱하는 데 특화되어 있으며 DETAIL 모드에서 CTE 임시 테이블의 상세 흐름을 분석함."
        },
        {
            "no": 9,
            "program": "sql_v12_full_emrput.py",
            "params": "<소스디렉토리>\n[--mode <SIMPLE/DETAIL>]\n[--db]\n[--conf <경로>]",
            "param_desc": "(sql_lng_015_emrput과 동일)\n- <소스디렉토리>: 분석 대상 디렉토리\n- --mode: 분석 모드 (SIMPLE 또는 DETAIL)",
            "opt_desc": "(sql_lng_015_emrput과 동일)",
            "features": "리니지 분석 기능과 함께, 실제 분석된 쿼리의 원본 텍스트를 담은 별도의 쿼리 텍스트 테이블({리니지테이블명}_query_text)을 자동으로 생성하고 적재(TRUNCATE 후 INSERT)하는 기능이 통합된 완성 버전."
        },
        {
            "no": 10,
            "program": "sql_v12_full_new.py",
            "params": "<검색대상_디렉토리>\n<경로포함_CSV파일명>\n[--mode <SIMPLE/DETAIL>]\n[--db]\n[--conf <경로>]",
            "param_desc": "- <검색대상_디렉토리>: 소스 파일 탐색 디렉토리 (필수)\n- <경로포함_CSV파일명>: table_name, column_name 헤더를 가진 기준 CSV 파일 경로 (필수)\n- --mode: 분석 모드 (SIMPLE 또는 DETAIL)\n- --db: DB 적재 활성화 여부\n- --conf: mysql.conf 경로",
            "opt_desc": "- <경로포함_CSV파일명>: 매칭 검토 기준이 되는 CSV 경로 (동적으로 읽어 in/ 폴더에 임시 저장)\n- --mode, --db, --conf는 기존 리니지 분석과 동일",
            "features": "sql_v12_full_emrput.py에 테이블/칼럼 매칭 기능을 결합함. 입력 CSV의 table_name 및 column_name과 일치하는 쿼리를 검색하여 칼럼매칭 결과 파일 및 DB 테이블({PROGRAM_NAME}_{last_dir}_{mode}_col_match)을 생성함."
        },
        {
            "no": 11,
            "program": "sql_v12_full_new_02.py",
            "params": "<검색대상_디렉토리>\n<경로포함_CSV파일명>\n[--mode <SIMPLE/DETAIL>]\n[--db]\n[--conf <경로>]",
            "param_desc": "- <경로포함_CSV파일명>: tbl_name, column_name, tobe_enc_key, tobe_enc_rsn 헤더를 가진 기준 CSV 파일 경로 (필수)\n(나머지는 sql_v12_full_new와 동일)",
            "opt_desc": "- <경로포함_CSV파일명>: 암호화 키 정보가 기재된 CSV 경로",
            "features": "sql_v12_full_new.py에서 입력 CSV의 헤더명(tbl_name)을 변경하고, 결과 파일/테이블에 암호화 키 및 사유 정보(tobe_enc_key, tobe_enc_rsn)가 함께 적재 및 출력되도록 기능을 보완함."
        },
        {
            "no": 12,
            "program": "sql_v12_full_new_03_local.py",
            "params": "<분석대상_디렉토리>\n<스키마.검색기준테이블>\n[--db]\n[--conf <경로>]",
            "param_desc": "- <분석대상_디렉토리>: 소스 파일 탐색 디렉토리 (필수)\n- <스키마.검색기준테이블>: 로컬 CSV 대신 검색 기준으로 사용하는 DB 내 테이블 (필수)\n- --db: DB 적재 활성화 여부\n- --conf: mysql.conf 경로",
            "opt_desc": "- <스키마.검색기준테이블>: 기준 정보가 담긴 DB 스키마/테이블 지정 (상시 DB 연동 동작)",
            "features": "기준 정보를 로컬 CSV 대신 DB 검색기준 테이블에서 조회하여 동작하며, tbl_name과 cols (복합 칼럼 정보, 예: col_01:k1,col_bb:k2)를 파싱하여 개별 칼럼 단위 매칭을 수행함. 매칭이 없는 행도 NULL로 보존하는 정합성이 추가됨."
        }
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "소스 프로그램 분석 요약"

    # 격자선(GridLines) 표시 설정
    ws.views.sheetView[0].showGridLines = True

    # 1. 스타일 설정
    font_title = Font(name="Malgun Gothic", size=16, bold=True, color="1F4E78")
    font_header = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Malgun Gothic", size=10)
    font_bold = Font(name="Malgun Gothic", size=10, bold=True)
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    border_thin = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 2. 타이틀 작성 (A2:F3 병합)
    ws.merge_cells("A2:F3")
    title_cell = ws["A2"]
    title_cell.value = "  ■ 소스 프로그램 매개변수 및 주요 기능 분석 요약"
    title_cell.font = font_title
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # 3. 헤더 작성 (5행)
    headers = ["No", "프로그램명", "파라미터", "파라미터 내용", "옵션 설명", "주요 기능"]
    header_row = 5
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    ws.row_dimensions[header_row].height = 30

    # 4. 데이터 본문 작성 (6행부터)
    start_row = 6
    for idx, row_data in enumerate(data):
        current_row = start_row + idx
        fill_color = fill_zebra if idx % 2 == 1 else fill_white

        # No
        c_no = ws.cell(row=current_row, column=1, value=row_data["no"])
        c_no.alignment = align_center
        
        # 프로그램명
        c_prog = ws.cell(row=current_row, column=2, value=row_data["program"])
        c_prog.alignment = align_left
        c_prog.font = font_bold
        
        # 파라미터
        c_params = ws.cell(row=current_row, column=3, value=row_data["params"])
        c_params.alignment = align_center
        
        # 파라미터 내용
        c_desc = ws.cell(row=current_row, column=4, value=row_data["param_desc"])
        c_desc.alignment = align_left
        
        # 옵션 설명
        c_opt = ws.cell(row=current_row, column=5, value=row_data["opt_desc"])
        c_opt.alignment = align_left
        
        # 주요 기능
        c_feat = ws.cell(row=current_row, column=6, value=row_data["features"])
        c_feat.alignment = align_left

        # 모든 셀에 폰트, 테두리, 채우기 적용
        for col_idx in range(1, 7):
            cell = ws.cell(row=current_row, column=col_idx)
            if col_idx != 2: # 프로그램명 굵은 폰트 유지
                cell.font = font_body
            cell.border = border_thin
            cell.fill = fill_color

        # 행 높이 설정 (가독성을 위해 넉넉하게 지정)
        ws.row_dimensions[current_row].height = 100

    # 5. 열 너비 명시적 지정
    col_widths = {
        "A": 6,   # No
        "B": 28,  # 프로그램명
        "C": 30,  # 파라미터
        "D": 45,  # 파라미터 내용
        "E": 45,  # 옵션 설명
        "F": 65   # 주요 기능
    }
    
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 파일 저장
    output_filename = "소스프로그램_분석_요약.xlsx"
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_filename)
    wb.save(output_path)
    print(f"[SUCCESS] 엑셀 파일이 정상적으로 생성되었습니다:\n -> {output_path}")
    return output_path

if __name__ == "__main__":
    create_excel()
